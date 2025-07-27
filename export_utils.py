import io
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
import streamlit as st

# Excel Export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    st.warning("⚠️ openpyxl tidak tersedia. Install dengan: pip install openpyxl")

# PDF Export  
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor, black, white, red, orange, green
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    st.warning("⚠️ reportlab tidak tersedia. Install dengan: pip install reportlab")

class ExcelExporter:
    def __init__(self):
        self.wb = None
        
    def create_excel_report(self, instansi_list: List, overlap_analysis: Dict[str, Any]) -> io.BytesIO:
        """Create comprehensive Excel report"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl tidak tersedia")
            
        self.wb = Workbook()
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Create sheets
        self._create_summary_sheet(overlap_analysis)
        self._create_instansi_sheet(instansi_list)
        self._create_overlap_sheet(overlap_analysis)
        self._create_recommendations_sheet(overlap_analysis)
        
        # Save to buffer
        excel_buffer = io.BytesIO()
        self.wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _create_summary_sheet(self, overlap_analysis: Dict[str, Any]):
        """Create summary sheet"""
        ws = self.wb.create_sheet("📊 Ringkasan Eksekutif")
        
        # Header styling
        header_font = Font(bold=True, size=16, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        
        # Title
        ws['A1'] = "LAPORAN ANALISIS TUMPANG TINDIH TUGAS INSTANSI PEMERINTAH"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:F1')
        
        # Metadata
        ws['A3'] = "👤 Dibuat oleh:"
        ws['B3'] = "itbahmad"
        ws['A4'] = "📅 Tanggal:"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = "🔧 Sistem:"
        ws['B5'] = "AI-Powered Government Overlap Analysis"
        
        # Executive Summary
        if 'ringkasan_eksekutif' in overlap_analysis:
            ws['A7'] = "📋 RINGKASAN EKSEKUTIF"
            ws['A7'].font = Font(bold=True, size=14)
            ws['A8'] = overlap_analysis['ringkasan_eksekutif']
            ws.merge_cells('A8:F10')
            ws['A8'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Key Metrics
        if 'metrik_overlap' in overlap_analysis:
            metrik = overlap_analysis['metrik_overlap']
            
            ws['A12'] = "📈 METRIK UTAMA"
            ws['A12'].font = Font(bold=True, size=14)
            
            metrics_data = [
                ["Metrik", "Nilai", "Status"],
                ["Total Tumpang Tindih", metrik.get('total_overlap_ditemukan', 0), "🔍"],
                ["Overlap Prioritas Tinggi", metrik.get('overlap_tinggi', 0), "🔴"],
                ["Overlap Prioritas Sedang", metrik.get('overlap_sedang', 0), "🟡"],
                ["Overlap Prioritas Rendah", metrik.get('overlap_rendah', 0), "🟢"],
                ["Efisiensi Potensial", metrik.get('efisiensi_potensial', '0%'), "💰"]
            ]
            
            for row_idx, row_data in enumerate(metrics_data, 13):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 13:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
        
        self._apply_table_styling(ws, 'A13:C18')
    
    def _create_instansi_sheet(self, instansi_list: List):
        """Create instansi data sheet"""
        ws = self.wb.create_sheet("🏢 Data Instansi")
        
        # Headers
        headers = [
            "No", "Nama Instansi", "Dokumen Sumber", "Jumlah Tugas Pokok", 
            "Jumlah Fungsi", "Jumlah Program", "Jumlah Kegiatan", "Total Items"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2a5298", end_color="2a5298", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        for row_idx, instansi in enumerate(instansi_list, 2):
            total_items = (len(instansi.tugas_pokok) + len(instansi.fungsi) + 
                          len(instansi.program) + len(instansi.kegiatan))
            
            row_data = [
                row_idx - 1,
                instansi.nama,
                ', '.join(instansi.dokumen_sumber),
                len(instansi.tugas_pokok),
                len(instansi.fungsi),
                len(instansi.program),
                len(instansi.kegiatan),
                total_items
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_overlap_sheet(self, overlap_analysis: Dict[str, Any]):
        """Create overlap analysis sheet"""
        ws = self.wb.create_sheet("🔍 Analisis Tumpang Tindih")
        
        # Headers
        headers = [
            "No", "Kategori", "Deskripsi", "Instansi Terlibat", 
            "Tingkat Overlap", "Dampak Potensial", "Estimasi Pemborosan"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        
        # Data
        if 'tumpang_tindih' in overlap_analysis:
            for row_idx, overlap in enumerate(overlap_analysis['tumpang_tindih'], 2):
                row_data = [
                    row_idx - 1,
                    overlap.get('kategori', '').replace('_', ' ').title(),
                    overlap.get('deskripsi', ''),
                    ', '.join(overlap.get('instansi_terlibat', [])),
                    overlap.get('tingkat_overlap', '').title(),
                    overlap.get('dampak_potensial', ''),
                    overlap.get('estimasi_pemborosan_anggaran', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Color coding based on overlap level
                    tingkat = overlap.get('tingkat_overlap', '').lower()
                    if tingkat == 'tinggi':
                        cell.fill = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
                    elif tingkat == 'sedang':
                        cell.fill = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
                    elif tingkat == 'rendah':
                        cell.fill = PatternFill(start_color="e8f5e8", end_color="e8f5e8", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_recommendations_sheet(self, overlap_analysis: Dict[str, Any]):
        """Create recommendations sheet"""
        ws = self.wb.create_sheet("💡 Rekomendasi")
        
        # Headers
        headers = [
            "No", "Prioritas", "Aksi", "Instansi Pelaksana", 
            "Timeline", "Benefit Estimasi"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2a5298", end_color="2a5298", fill_type="solid")
        
        # Data
        if 'rekomendasi' in overlap_analysis:
            for row_idx, rec in enumerate(overlap_analysis['rekomendasi'], 2):
                row_data = [
                    row_idx - 1,
                    rec.get('prioritas', '').title(),
                    rec.get('aksi', ''),
                    rec.get('instansi_pelaksana', ''),
                    rec.get('timeline', ''),
                    rec.get('benefit_estimasi', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Priority color coding
                    prioritas = rec.get('prioritas', '').lower()
                    if prioritas == 'tinggi':
                        cell.fill = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")
                    elif prioritas == 'sedang':
                        cell.fill = PatternFill(start_color="ffe0b2", end_color="ffe0b2", fill_type="solid")
                    elif prioritas == 'rendah':
                        cell.fill = PatternFill(start_color="dcedc8", end_color="dcedc8", fill_type="solid")
        
        # Adjust row heights for wrapped text
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 60
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_table_styling(self, ws, range_string):
        """Apply table styling to a range"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws[range_string]:
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

class PDFExporter:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom styles for PDF"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Title'],
            fontSize=18,
            spaceAfter=30,
            textColor=HexColor('#1e3c72'),
            alignment=1  # Center
        ))
        
        # Header style
        self.styles.add(ParagraphStyle(
            name='CustomHeader',
            parent=self.styles['Heading1'],
            fontSize=14,
            spaceAfter=12,
            textColor=HexColor('#2a5298'),
            leftIndent=0
        ))
        
        # Subheader style
        self.styles.add(ParagraphStyle(
            name='CustomSubHeader',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=8,
            textColor=HexColor('#1e3c72'),
            leftIndent=10
        ))
    
    def create_pdf_report(self, instansi_list: List, overlap_analysis: Dict[str, Any]) -> io.BytesIO:
        """Create comprehensive PDF report"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab tidak tersedia")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            rightMargin=72, 
            leftMargin=72,
            topMargin=72, 
            bottomMargin=18
        )
        
        # Build story
        story = []
        
        # Title page
        story.extend(self._create_title_page())
        story.append(PageBreak())
        
        # Executive summary
        story.extend(self._create_executive_summary(overlap_analysis))
        story.append(PageBreak())
        
        # Instansi data
        story.extend(self._create_instansi_section(instansi_list))
        story.append(PageBreak())
        
        # Overlap analysis
        story.extend(self._create_overlap_section(overlap_analysis))
        story.append(PageBreak())
        
        # Recommendations
        story.extend(self._create_recommendations_section(overlap_analysis))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
    
    def _create_title_page(self):
        """Create title page"""
        story = []
        
        # Main title
        title = Paragraph(
            "LAPORAN ANALISIS TUMPANG TINDIH TUGAS INSTANSI PEMERINTAH",
            self.styles['CustomTitle']
        )
        story.append(title)
        story.append(Spacer(1, 40))
        
        # Subtitle
        subtitle = Paragraph(
            "Sistem Identifikasi Otomatis Duplikasi Tugas, Fungsi, dan Program Antar Instansi",
            self.styles['Heading2']
        )
        story.append(subtitle)
        story.append(Spacer(1, 60))
        
        # Metadata table
        metadata = [
            ["Dibuat oleh:", "itbahmad"],
            ["Tanggal:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Sistem:", "AI-Powered Government Overlap Analysis"],
            ["Powered by:", "Google Gemini Pro API"]
        ]
        
        metadata_table = Table(metadata, colWidths=[2*inch, 3*inch])
        metadata_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey)
        ]))
        
        story.append(metadata_table)
        story.append(Spacer(1, 40))
        
        # Disclaimer
        disclaimer = Paragraph(
            "<i>Laporan ini dibuat secara otomatis menggunakan teknologi AI untuk membantu "
            "identifikasi tumpang tindih tugas antar instansi pemerintah. "
            "Hasil analisis perlu diverifikasi lebih lanjut oleh pihak yang berwenang.</i>",
            self.styles['Normal']
        )
        story.append(disclaimer)
        
        return story
    
    def _create_executive_summary(self, overlap_analysis):
        """Create executive summary section"""
        story = []
        
        story.append(Paragraph("RINGKASAN EKSEKUTIF", self.styles['CustomHeader']))
        
        if 'ringkasan_eksekutif' in overlap_analysis:
            summary = Paragraph(overlap_analysis['ringkasan_eksekutif'], self.styles['Normal'])
            story.append(summary)
            story.append(Spacer(1, 20))
        
        # Key metrics
        if 'metrik_overlap' in overlap_analysis:
            story.append(Paragraph("Metrik Utama", self.styles['CustomSubHeader']))
            
            metrik = overlap_analysis['metrik_overlap']
            metrics_data = [
                ["Metrik", "Nilai"],
                ["Total Tumpang Tindih Ditemukan", str(metrik.get('total_overlap_ditemukan', 0))],
                ["Overlap Prioritas Tinggi", str(metrik.get('overlap_tinggi', 0))],
                ["Overlap Prioritas Sedang", str(metrik.get('overlap_sedang', 0))],
                ["Overlap Prioritas Rendah", str(metrik.get('overlap_rendah', 0))],
                ["Efisiensi Potensial", str(metrik.get('efisiensi_potensial', '0%'))]
            ]
            
            metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(metrics_table)
        
        return story
    
    def _create_instansi_section(self, instansi_list):
        """Create instansi data section"""
        story = []
        
        story.append(Paragraph("DATA INSTANSI", self.styles['CustomHeader']))
        
        # Summary table
        instansi_data = [["No", "Nama Instansi", "Jumlah Dokumen", "Total Items"]]
        
        for idx, instansi in enumerate(instansi_list, 1):
            total_items = (len(instansi.tugas_pokok) + len(instansi.fungsi) + 
                          len(instansi.program) + len(instansi.kegiatan))
            
            instansi_data.append([
                str(idx),
                instansi.nama,
                str(len(instansi.dokumen_sumber)),
                str(total_items)
            ])
        
        instansi_table = Table(instansi_data, colWidths=[0.5*inch, 3*inch, 1*inch, 1*inch])
        instansi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2a5298')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        
        story.append(instansi_table)
        story.append(Spacer(1, 20))
        
        # Detailed breakdown
        for idx, instansi in enumerate(instansi_list, 1):
            story.append(Paragraph(f"{idx}. {instansi.nama}", self.styles['CustomSubHeader']))
            
            details = f"""
            <b>Dokumen Sumber:</b> {', '.join(instansi.dokumen_sumber)}<br/>
            <b>Tugas Pokok:</b> {len(instansi.tugas_pokok)} items<br/>
            <b>Fungsi:</b> {len(instansi.fungsi)} items<br/>
            <b>Program:</b> {len(instansi.program)} items<br/>
            <b>Kegiatan:</b> {len(instansi.kegiatan)} items
            """
            
            story.append(Paragraph(details, self.styles['Normal']))
            story.append(Spacer(1, 10))
        
        return story
    
    def _create_overlap_section(self, overlap_analysis):
        """Create overlap analysis section"""
        story = []
        
        story.append(Paragraph("ANALISIS TUMPANG TINDIH", self.styles['CustomHeader']))
        
        if 'tumpang_tindih' in overlap_analysis:
            overlaps = overlap_analysis['tumpang_tindih']
            
            for idx, overlap in enumerate(overlaps, 1):
                # Priority indicator
                tingkat = overlap.get('tingkat_overlap', '').upper()
                priority_color = {
                    'TINGGI': 'red',
                    'SEDANG': 'orange', 
                    'RENDAH': 'green'
                }.get(tingkat, 'black')
                
                title = f"<font color='{priority_color}'>{idx}. {overlap.get('kategori', '').replace('_', ' ').title()} - {tingkat}</font>"
                story.append(Paragraph(title, self.styles['CustomSubHeader']))
                
                # Details
                details = f"""
                <b>Deskripsi:</b> {overlap.get('deskripsi', '')}<br/>
                <b>Instansi Terlibat:</b> {', '.join(overlap.get('instansi_terlibat', []))}<br/>
                <b>Dampak Potensial:</b> {overlap.get('dampak_potensial', '')}<br/>
                <b>Estimasi Pemborosan:</b> {overlap.get('estimasi_pemborosan_anggaran', 'Tidak tersedia')}
                """
                
                story.append(Paragraph(details, self.styles['Normal']))
                story.append(Spacer(1, 15))
        
        return story
    
    def _create_recommendations_section(self, overlap_analysis):
        """Create recommendations section"""
        story = []
        
        story.append(Paragraph("REKOMENDASI STRATEGIS", self.styles['CustomHeader']))
        
        if 'rekomendasi' in overlap_analysis:
            recommendations = overlap_analysis['rekomendasi']
            
            for idx, rec in enumerate(recommendations, 1):
                # Priority indicator
                prioritas = rec.get('prioritas', '').upper()
                priority_color = {
                    'TINGGI': 'red',
                    'SEDANG': 'orange',
                    'RENDAH': 'green'
                }.get(prioritas, 'black')
                
                title = f"<font color='{priority_color}'>{idx}. Prioritas {prioritas}</font>"
                story.append(Paragraph(title, self.styles['CustomSubHeader']))
                
                # Details
                details = f"""
                <b>Aksi:</b> {rec.get('aksi', '')}<br/>
                <b>Instansi Pelaksana:</b> {rec.get('instansi_pelaksana', '')}<br/>
                <b>Timeline:</b> {rec.get('timeline', '')}<br/>
                <b>Benefit Estimasi:</b> {rec.get('benefit_estimasi', '')}
                """
                
                story.append(Paragraph(details, self.styles['Normal']))
                story.append(Spacer(1, 15))
        
        return story

# Export functions for use in main app
def create_excel_report(instansi_list: List, overlap_analysis: Dict[str, Any]):
    """Create and download Excel report"""
    try:
        exporter = ExcelExporter()
        excel_buffer = exporter.create_excel_report(instansi_list, overlap_analysis)
        
        filename = f"laporan_tumpang_tindih_itbahmad_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_buffer.getvalue(),
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="excel_download"
        )
        
        st.success(f"✅ Excel report siap didownload: {filename}")
        
    except ImportError as e:
        st.error(f"❌ Error: {e}. Install dengan: pip install openpyxl")
    except Exception as e:
        st.error(f"❌ Error membuat Excel report: {e}")
        st.exception(e)

def create_pdf_report(instansi_list: List, overlap_analysis: Dict[str, Any]):
    """Create and download PDF report"""
    try:
        exporter = PDFExporter()
        pdf_buffer = exporter.create_pdf_report(instansi_list, overlap_analysis)
        
        filename = f"laporan_tumpang_tindih_itbahmad_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        st.download_button(
            label="📄 Download PDF Report",
            data=pdf_buffer.getvalue(),
            file_name=filename,
            mime="application/pdf",
            key="pdf_download"
        )
        
        st.success(f"✅ PDF report siap didownload: {filename}")
        
    except ImportError as e:
        st.error(f"❌ Error: {e}. Install dengan: pip install reportlab")
    except Exception as e:
        st.error(f"❌ Error membuat PDF report: {e}")
        st.exception(e)